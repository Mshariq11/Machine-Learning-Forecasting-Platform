"""
=========================================================
Excel Reporting Module
=========================================================

Generate professional Excel reports for
Store Sales Forecasting.

Phase 1
--------
Sheets:
- Summary
- Forecast
- Metadata

Author : Shariq Zia
Project: Store Sales Forecasting
"""

from datetime import datetime
from pathlib import Path
from typing import Dict

import numpy as np
import pandas as pd

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.chart import (
    LineChart,
    Reference
)


# =========================================================
# STYLES
# =========================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=11
)

TITLE_FONT = Font(
    bold=True,
    size=16
)

SUBTITLE_FONT = Font(
    bold=True,
    size=12
)


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def auto_adjust_columns(ws):
    """
    Automatically resize worksheet columns.
    """

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 3


def style_header(row):
    """
    Apply header style.
    """

    for cell in row:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center"
        )


# =========================================================
# WORKBOOK
# =========================================================

def create_workbook():
    """
    Create workbook.
    """

    wb = Workbook()

    wb.remove(
        wb.active
    )

    return wb


# =========================================================
# SUMMARY SHEET
# =========================================================

def create_summary_sheet(
    wb: Workbook,
    forecast: pd.DataFrame,
    metadata: Dict,
    company_name: str,
    report_title: str
):
    """
    Executive summary.
    """

    ws = wb.create_sheet("Summary")

    ws["A1"] = company_name
    ws["A1"].font = TITLE_FONT

    ws["A2"] = report_title
    ws["A2"].font = SUBTITLE_FONT

    total_sales = forecast["sales"].sum()

    avg_sales = forecast["sales"].mean()

    max_sales = forecast["sales"].max()

    min_sales = forecast["sales"].min()

    horizon = (
        forecast["date"].nunique()
        if "date" in forecast.columns
        else len(forecast)
    )

    rows = [

        ("Model",
         metadata.get("model", "Unknown")),

        ("Training Date",
         str(metadata.get("training_date", ""))),

        ("Target",
         metadata.get("target", "")),

        ("Forecast Horizon",
         horizon),

        ("Forecast Records",
         len(forecast)),

        ("Total Forecast Sales",
         total_sales),

        ("Average Forecast",
         avg_sales),

        ("Maximum Forecast",
         max_sales),

        ("Minimum Forecast",
         min_sales)

    ]

    start = 5

    ws.cell(
        row=start,
        column=1
    ).value = "Metric"

    ws.cell(
        row=start,
        column=2
    ).value = "Value"

    style_header(
        ws[start]
    )

    for i, row in enumerate(rows, start + 1):

        ws.cell(
            row=i,
            column=1
        ).value = row[0]

        ws.cell(
            row=i,
            column=2
        ).value = row[1]

    auto_adjust_columns(ws)


# =========================================================
# FORECAST SHEET
# =========================================================

def create_forecast_sheet(
    wb: Workbook,
    forecast: pd.DataFrame
):
    """
    Forecast data.
    """

    ws = wb.create_sheet(
        "Forecast"
    )

    headers = list(
        forecast.columns
    )

    for col, header in enumerate(
        headers,
        1
    ):

        ws.cell(
            row=1,
            column=col
        ).value = header

    style_header(
        ws[1]
    )

    for row in forecast.itertuples(
        index=False
    ):

        ws.append(
            list(row)
        )

    ws.freeze_panes = "A2"

    auto_adjust_columns(ws)


# =========================================================
# METADATA SHEET
# =========================================================

def create_metadata_sheet(
    wb: Workbook,
    metadata: Dict
):
    """
    Metadata sheet.
    """

    ws = wb.create_sheet(
        "Metadata"
    )

    ws["A1"] = "Item"
    ws["B1"] = "Value"

    style_header(
        ws[1]
    )

    row = 2

    for key, value in metadata.items():

        ws.cell(
            row=row,
            column=1
        ).value = key

        ws.cell(
            row=row,
            column=2
        ).value = str(value)

        row += 1

    auto_adjust_columns(ws)


# =========================================================
# SAVE REPORT
# =========================================================

def save_report(
    wb: Workbook,
    output_path: Path
):
    """
    Save workbook.
    """

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(
        output_path
    )


# =========================================================
# MASTER REPORT
# =========================================================

def create_excel_report(
    forecast: pd.DataFrame,
    metadata: Dict,
    output_path: Path,
    company_name: str = "ABC Retail",
    report_title: str = "Demand Forecast Report"
):
    """
    Generate complete Excel report.
    """

    print("\n" + "=" * 70)
    print("GENERATING EXCEL REPORT")
    print("=" * 70)

    wb = create_workbook()

    create_summary_sheet(
        wb,
        forecast,
        metadata,
        company_name,
        report_title
    )

    create_forecast_sheet(
        wb,
        forecast
    )

    create_metadata_sheet(
        wb,
        metadata
    )

    save_report(
        wb,
        output_path
    )

    print(
        f"\nReport saved to:\n{output_path}"
    )

    print("\n" + "=" * 70)
    print("REPORT GENERATION COMPLETED")
    print("=" * 70)

# =========================================================
# EXCEL STYLING
# =========================================================

def style_header(ws):
    """
    Style worksheet header.
    """

    fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in ws[1]:

        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def auto_adjust_columns(ws):
    """
    Automatically adjust column widths.
    """

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 3, 40)


# =========================================================
# COMPANY INFORMATION
# =========================================================

def add_company_information(
    workbook
):
    """
    Create metadata worksheet.
    """

    ws = workbook.create_sheet(
        "Metadata"
    )

    rows = [

        ["Company", "ABC Retail Corporation"],

        ["Department", "Demand Planning"],

        ["Forecast Generated", datetime.now()],

        ["Forecast Horizon", "16 Days"],

        ["Model", "XGBoost"],

        ["Prepared By", "Shariq Zia"],

        ["Target Variable", "Sales"],

        ["Project", "Store Sales Forecasting"],

        ["Version", "1.0"]

    ]

    for row in rows:
        ws.append(row)

    style_header(ws)
    auto_adjust_columns(ws)


# =========================================================
# SUMMARY SHEET
# =========================================================

def create_summary_sheet(
    workbook,
    forecast: pd.DataFrame
):
    """
    Executive summary.
    """

    ws = workbook.create_sheet(
        "Summary"
    )

    total_sales = forecast[
        "Forecast"
    ].sum()

    avg_sales = forecast[
        "Forecast"
    ].mean()

    max_sales = forecast[
        "Forecast"
    ].max()

    min_sales = forecast[
        "Forecast"
    ].min()

    rows = [

        ["Metric", "Value"],

        ["Forecast Records", len(forecast)],

        ["Total Forecast", round(total_sales,2)],

        ["Average Forecast", round(avg_sales,2)],

        ["Maximum Forecast", round(max_sales,2)],

        ["Minimum Forecast", round(min_sales,2)]

    ]

    for row in rows:
        ws.append(row)

    style_header(ws)
    auto_adjust_columns(ws)


# =========================================================
# FORECAST SHEET
# =========================================================

def create_forecast_sheet(
    workbook,
    forecast: pd.DataFrame
):
    """
    Forecast table.
    """

    ws = workbook.create_sheet(
        "Forecast"
    )

    for row in dataframe_to_rows(
        forecast,
        index=False,
        header=True
    ):
        ws.append(row)

    style_header(ws)

    auto_adjust_columns(ws)


# =========================================================
# ACTUAL VS FORECAST
# =========================================================

def create_actual_vs_forecast_sheet(
    workbook,
    comparison: pd.DataFrame
):
    """
    Compare actual and predicted sales.
    """

    df = comparison.copy()

    df["Error"] = (

        df["Forecast"]

        -

        df["Actual"]

    )

    df["Absolute Error"] = (

        df["Error"]

        .abs()

    )

    df["APE (%)"] = np.where(

        df["Actual"] == 0,

        np.nan,

        (

            df["Absolute Error"]

            /

            df["Actual"]

        ) * 100

    )

    ws = workbook.create_sheet(
        "Actual vs Forecast"
    )

    for row in dataframe_to_rows(
        df,
        index=False,
        header=True
    ):
        ws.append(row)

    style_header(ws)

    auto_adjust_columns(ws)

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    error_col = None

    for cell in ws[1]:

        if cell.value == "APE (%)":

            error_col = cell.column_letter

            break

    if error_col:

        for row in range(2, ws.max_row + 1):

            cell = ws[f"{error_col}{row}"]

            if cell.value is None:
                continue

            if cell.value > 20:

                cell.fill = red_fill

            else:

                cell.fill = green_fill

# =========================================================
# STORE PERFORMANCE SHEET
# =========================================================

def create_store_performance_sheet(
    workbook,
    forecast: pd.DataFrame
):
    """
    Store level forecast summary.
    """

    if "store_nbr" not in forecast.columns:
        return

    summary = (
        forecast
        .groupby("store_nbr", as_index=False)
        .agg(
            Forecast=("Forecast", "sum")
        )
        .sort_values(
            "Forecast",
            ascending=False
        )
    )

    ws = workbook.create_sheet(
        "Store Performance"
    )

    for row in dataframe_to_rows(
        summary,
        index=False,
        header=True
    ):
        ws.append(row)

    style_header(ws)
    auto_adjust_columns(ws)


# =========================================================
# PRODUCT FAMILY SHEET
# =========================================================

def create_family_sheet(
    workbook,
    forecast: pd.DataFrame
):
    """
    Product family forecast.
    """

    if "family" not in forecast.columns:
        return

    summary = (
        forecast
        .groupby("family", as_index=False)
        .agg(
            Forecast=("Forecast", "sum")
        )
        .sort_values(
            "Forecast",
            ascending=False
        )
    )

    ws = workbook.create_sheet(
        "Product Families"
    )

    for row in dataframe_to_rows(
        summary,
        index=False,
        header=True
    ):
        ws.append(row)

    style_header(ws)
    auto_adjust_columns(ws)


# =========================================================
# INVENTORY PLANNING
# =========================================================

def create_inventory_sheet(
    workbook,
    forecast: pd.DataFrame
):
    """
    Inventory planning recommendations.
    """

    df = forecast.copy()

    df["Recommended Stock"] = (
        np.ceil(
            df["Forecast"] * 1.20
        )
    )

    df["Safety Stock"] = (
        np.ceil(
            df["Forecast"] * 0.20
        )
    )

    df["Reorder Point"] = (
        np.ceil(
            df["Forecast"] * 0.80
        )
    )

    columns = [
        c
        for c in [
            "store_nbr",
            "family",
            "Forecast",
            "Recommended Stock",
            "Safety Stock",
            "Reorder Point"
        ]
        if c in df.columns
    ]

    ws = workbook.create_sheet(
        "Inventory Planning"
    )

    for row in dataframe_to_rows(
        df[columns],
        index=False,
        header=True
    ):
        ws.append(row)

    style_header(ws)
    auto_adjust_columns(ws)


# =========================================================
# CHARTS
# =========================================================

def create_chart_sheet(
    workbook,
    forecast: pd.DataFrame
):
    """
    Create forecast charts.
    """

    ws = workbook.create_sheet(
        "Charts"
    )

    if "date" not in forecast.columns:
        return

    daily = (
        forecast
        .groupby("date", as_index=False)
        .agg(
            Forecast=("Forecast", "sum")
        )
    )

    for row in dataframe_to_rows(
        daily,
        index=False,
        header=True
    ):
        ws.append(row)

    style_header(ws)

    chart = LineChart()

    data = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=len(daily)+1
    )

    categories = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=len(daily)+1
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(
        categories
    )

    chart.title = "Forecast Trend"

    chart.y_axis.title = "Sales"

    chart.x_axis.title = "Date"

    ws.add_chart(
        chart,
        "E2"
    )

    auto_adjust_columns(ws)


# =========================================================
# EXPORT REPORT
# =========================================================

def create_excel_report(
    forecast: pd.DataFrame,
    output_path: Path,
    actuals: pd.DataFrame | None = None
):
    """
    Generate complete Excel report.
    """

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    create_summary_sheet(
        workbook,
        forecast
    )

    create_forecast_sheet(
        workbook,
        forecast
    )

    create_store_performance_sheet(
        workbook,
        forecast
    )

    create_family_sheet(
        workbook,
        forecast
    )

    create_inventory_sheet(
        workbook,
        forecast
    )

    create_chart_sheet(
        workbook,
        forecast
    )

    if actuals is not None:

        comparison = (
            actuals.merge(
                forecast,
                on="id",
                how="inner"
            )
        )

        comparison = comparison.rename(
            columns={
                "sales": "Actual"
            }
        )

        create_actual_vs_forecast_sheet(
            workbook,
            comparison
        )

    add_company_information(
        workbook
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook.save(
        output_path
    )

    print("\n" + "=" * 70)
    print("EXCEL REPORT GENERATED")
    print("=" * 70)
    print(output_path)
    print("=" * 70)

# =========================================================
# MASTER FUNCTION
# =========================================================

def run_excel_report(
    forecast: pd.DataFrame,
    output_path: Path,
    actuals: pd.DataFrame | None = None
):
    """
    Generate complete business report.
    """

    print("\n" + "=" * 70)
    print("GENERATING EXCEL REPORT")
    print("=" * 70)

    create_excel_report(
        forecast=forecast,
        output_path=output_path,
        actuals=actuals
    )

    print("\n" + "=" * 70)
    print("REPORT COMPLETED")
    print("=" * 70)